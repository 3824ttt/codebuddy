"""
读者荐读系统 - 后端 API
Flask + SQLite + Excel导入
"""
import json
import os
import sqlite3
import io
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, g, send_from_directory, send_file
from classifier import classify_book, classify_by_rules, AI_CONFIG as AI_CLASSIFIER_CONFIG, recommend_majors

app = Flask(__name__, static_folder='../frontend', static_url_path='')
DATABASE = os.path.join(os.path.dirname(__file__), 'library.db')
ADMIN_PASSWORD = 'admin123'

# ========== 阿里云 ISBN API 配置 ==========
# AppCode 获取：https://market.aliyun.com/products/57126001/cmapi00037193.html
ALICLOUD_APPCODE = '65725bb8242040df86e0a0e54cca9714'

# ========== Excel 支持 ==========
try:
    import openpyxl
except ImportError:
    openpyxl = None


# ==================== 数据库 ====================
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db:
        db.close()


def init_db():
    db = sqlite3.connect(DATABASE)
    db.execute("PRAGMA journal_mode=WAL")
    db.executescript('''
        CREATE TABLE IF NOT EXISTS books (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            subtitle TEXT DEFAULT '',
            volume_name TEXT DEFAULT '',
            volume_no TEXT DEFAULT '',
            author TEXT DEFAULT '',
            series_name TEXT DEFAULT '',
            series_author TEXT DEFAULT '',
            publisher TEXT DEFAULT '',
            isbn TEXT DEFAULT '',
            price REAL DEFAULT 0.0,
            pub_year TEXT DEFAULT '',
            pages TEXT DEFAULT '',
            category TEXT DEFAULT '',
            clc_number TEXT DEFAULT '',
            keywords TEXT DEFAULT '',
            description TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            cover_url TEXT DEFAULT '',
            click_count INTEGER DEFAULT 0,
            status TEXT DEFAULT 'pending',
            recommended_major TEXT DEFAULT '',
            recommended_subject TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS click_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            book_id INTEGER NOT NULL,
            reader_ip TEXT DEFAULT '',
            reader_id TEXT DEFAULT '',
            clicked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (book_id) REFERENCES books(id)
        );

        CREATE TABLE IF NOT EXISTS config (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor TEXT DEFAULT '',
            contact TEXT DEFAULT '',
            note TEXT DEFAULT '',
            book_ids TEXT DEFAULT '[]',
            total_amount REAL DEFAULT 0.0,
            status TEXT DEFAULT 'draft',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER NOT NULL,
            book_id INTEGER NOT NULL,
            quantity INTEGER DEFAULT 1,
            FOREIGN KEY (order_id) REFERENCES orders(id),
            FOREIGN KEY (book_id) REFERENCES books(id)
        );

        CREATE TABLE IF NOT EXISTS isbn_cache (
            isbn TEXT PRIMARY KEY,
            title TEXT DEFAULT '',
            subtitle TEXT DEFAULT '',
            author TEXT DEFAULT '',
            publisher TEXT DEFAULT '',
            pub_year TEXT DEFAULT '',
            pages TEXT DEFAULT '',
            description TEXT DEFAULT '',
            cover_url TEXT DEFAULT '',
            fetched_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        INSERT OR IGNORE INTO config (key, value) VALUES ('click_threshold', '10');
        INSERT OR IGNORE INTO config (key, value) VALUES ('click_cooldown_hours', '24');
    ''')
    # 迁移：为旧数据库添加新字段
    for col in ['recommended_major', 'recommended_subject']:
        try:
            db.execute(f'ALTER TABLE books ADD COLUMN {col} TEXT DEFAULT \'\'')
        except sqlite3.OperationalError:
            pass
    db.commit()
    db.close()


# ==================== 工具 ====================
def get_config(key):
    db = get_db()
    row = db.execute('SELECT value FROM config WHERE key = ?', (key,)).fetchone()
    return row['value'] if row else None


def set_config(key, value):
    db = get_db()
    db.execute('INSERT OR REPLACE INTO config (key, value) VALUES (?, ?)', (key, str(value)))
    db.commit()


def get_client_id():
    ip = request.remote_addr or 'unknown'
    ua = request.headers.get('User-Agent', '')[:50]
    return f"{ip}|{ua}"


def can_click(book_id, reader_id):
    db = get_db()
    hours = int(get_config('click_cooldown_hours') or 24)
    since = datetime.now() - timedelta(hours=hours)
    row = db.execute(
        'SELECT id FROM click_records WHERE book_id=? AND reader_id=? AND clicked_at > ?',
        (book_id, reader_id, since)
    ).fetchone()
    return row is None


BOOK_FIELDS = [
    'title', 'subtitle', 'volume_name', 'volume_no', 'author',
    'series_name', 'series_author', 'publisher', 'isbn', 'price',
    'pub_year', 'pages', 'category', 'clc_number', 'keywords',
    'description', 'notes', 'cover_url'
]


def _safe_to_float(v):
    """安全转换为 float，非数值返回 0.0"""
    if v is None or (isinstance(v, str) and v.strip() == ''):
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _insert_book(db, b):
    """插入一本图书，返回 (inserted, skipped)"""
    # 标题为唯一必要字段
    title = (b.get('title') or '').strip()
    if not title:
        return (0, 1)

    isbn = (b.get('isbn') or '').strip()
    if isbn:
        exist = db.execute('SELECT id FROM books WHERE isbn = ?', (isbn,)).fetchone()
        if exist:
            return (0, 1)

    vals = []
    placeholders = []
    for f in BOOK_FIELDS:
        val = b.get(f, '')
        if val is None:
            val = ''
        if f == 'price':
            val = _safe_to_float(val)
        else:
            val = str(val).strip()
        vals.append(val)
        placeholders.append('?')

    try:
        db.execute(
            f'INSERT INTO books ({",".join(BOOK_FIELDS)}) VALUES ({",".join(placeholders)})',
            vals
        )
        return (1, 0)
    except Exception as e:
        # 插入失败（极少情况），返回跳过
        return (0, 1)


# ==================== 读者 API ====================
@app.route('/api/books', methods=['GET'])
def get_books():
    db = get_db()
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    publisher = request.args.get('publisher', '')
    subject = request.args.get('subject', '')
    major = request.args.get('major', '')
    sort = request.args.get('sort', 'click_count')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))

    where = ['1=1']
    params = []

    if search:
        where.append('(title LIKE ? OR author LIKE ? OR isbn LIKE ? OR keywords LIKE ? OR publisher LIKE ? OR recommended_major LIKE ? OR recommended_subject LIKE ?)')
        params.extend([f'%{search}%'] * 7)

    if category:
        where.append('category = ?')
        params.append(category)

    if publisher:
        where.append('publisher = ?')
        params.append(publisher)

    if subject:
        where.append('recommended_subject = ?')
        params.append(subject)

    if major:
        where.append('recommended_major LIKE ?')
        params.append(f'%{major}%')

    where_clause = ' AND '.join(where)

    count_row = db.execute(f'SELECT COUNT(*) FROM books WHERE {where_clause}', params).fetchone()
    total = count_row[0]

    order_map = {
        'click_count': 'click_count DESC',
        'newest': 'created_at DESC',
        'title': 'title ASC'
    }
    order = order_map.get(sort, 'click_count DESC')

    offset = (page - 1) * per_page
    rows = db.execute(
        f'SELECT * FROM books WHERE {where_clause} ORDER BY {order} LIMIT ? OFFSET ?',
        params + [per_page, offset]
    ).fetchall()

    threshold = int(get_config('click_threshold') or 10)
    books = []
    for row in rows:
        d = dict(row)
        d['progress'] = min(100, round(d['click_count'] / threshold * 100, 1)) if threshold > 0 else 100
        d['reach_threshold'] = d['click_count'] >= threshold
        books.append(d)

    cats = db.execute('SELECT DISTINCT category FROM books WHERE category != "" ORDER BY category').fetchall()
    categories = [c['category'] for c in cats]

    pubs = db.execute("SELECT DISTINCT publisher FROM books WHERE publisher != '' ORDER BY publisher").fetchall()
    publishers = [p['publisher'] for p in pubs]

    subjs = db.execute("SELECT DISTINCT recommended_subject FROM books WHERE recommended_subject != '' ORDER BY recommended_subject").fetchall()
    subjects = [s['recommended_subject'] for s in subjs]

    majs = db.execute("SELECT DISTINCT recommended_major FROM books WHERE recommended_major != '' ORDER BY recommended_major").fetchall()
    majors = []
    for m in majs:
        for part in m['recommended_major'].split('、'):
            part = part.strip()
            if part and part not in majors:
                majors.append(part)

    return jsonify({
        'books': books, 'total': total, 'page': page,
        'per_page': per_page, 'categories': categories,
        'publishers': publishers, 'subjects': subjects,
        'majors': sorted(majors), 'threshold': threshold
    })


@app.route('/api/books/<int:book_id>', methods=['GET'])
def get_book_detail(book_id):
    db = get_db()
    row = db.execute('SELECT * FROM books WHERE id = ?', (book_id,)).fetchone()
    if not row:
        return jsonify({'error': '图书不存在'}), 404
    threshold = int(get_config('click_threshold') or 10)
    d = dict(row)
    d['progress'] = min(100, round(d['click_count'] / threshold * 100, 1)) if threshold > 0 else 100
    d['reach_threshold'] = d['click_count'] >= threshold
    return jsonify(d)


@app.route('/api/books/<int:book_id>/click', methods=['POST'])
def click_book(book_id):
    db = get_db()
    row = db.execute('SELECT id, title, click_count FROM books WHERE id = ?', (book_id,)).fetchone()
    if not row:
        return jsonify({'error': '图书不存在'}), 404

    reader_id = get_client_id()
    if not can_click(book_id, reader_id):
        hours = get_config('click_cooldown_hours') or 24
        return jsonify({'error': f'您在 {hours} 小时内已推荐过该书，请稍后再来', 'cooldown': True}), 429

    db.execute(
        'INSERT INTO click_records (book_id, reader_ip, reader_id) VALUES (?, ?, ?)',
        (book_id, request.remote_addr or 'unknown', reader_id)
    )
    db.execute('UPDATE books SET click_count = click_count + 1, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (book_id,))
    threshold = int(get_config('click_threshold') or 10)
    new_count = row['click_count'] + 1
    if new_count >= threshold:
        db.execute("UPDATE books SET status = 'recommended', updated_at = CURRENT_TIMESTAMP WHERE id = ? AND status = 'pending'", (book_id,))
    db.commit()

    return jsonify({'success': True, 'click_count': new_count, 'reach_threshold': new_count >= threshold})


@app.route('/api/books/trending', methods=['GET'])
def trending_books():
    db = get_db()
    rows = db.execute('SELECT * FROM books ORDER BY click_count DESC LIMIT 10').fetchall()
    return jsonify([dict(r) for r in rows])


# ==================== 管理后台 ====================
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('X-Admin-Token', '')
        if token != ADMIN_PASSWORD:
            return jsonify({'error': '无权限访问'}), 401
        return f(*args, **kwargs)
    return decorated


@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    data = request.get_json()
    password = data.get('password', '')
    if password == ADMIN_PASSWORD:
        return jsonify({'token': ADMIN_PASSWORD, 'success': True})
    return jsonify({'error': '密码错误'}), 401


@app.route('/api/admin/books', methods=['GET'])
@admin_required
def admin_books():
    db = get_db()
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))

    where = ['1=1']
    params = []
    if search:
        where.append('(title LIKE ? OR isbn LIKE ? OR author LIKE ?)')
        params.extend([f'%{search}%'] * 3)
    if status:
        where.append('status = ?')
        params.append(status)

    where_clause = ' AND '.join(where)
    count = db.execute(f'SELECT COUNT(*) FROM books WHERE {where_clause}', params).fetchone()[0]
    rows = db.execute(
        f'SELECT * FROM books WHERE {where_clause} ORDER BY updated_at DESC LIMIT ? OFFSET ?',
        params + [per_page, (page - 1) * per_page]
    ).fetchall()

    threshold = int(get_config('click_threshold') or 10)
    books = []
    for r in rows:
        d = dict(r)
        d['progress'] = min(100, round(d['click_count'] / threshold * 100, 1)) if threshold > 0 else 100
        books.append(d)

    return jsonify({'books': books, 'total': count, 'page': page, 'per_page': per_page})


# ------- 导入（JSON + Excel） -------
@app.route('/api/admin/books/import', methods=['POST'])
@admin_required
def import_books_json():
    """批量导入图书 - JSON 格式"""
    data = request.get_json()
    books = data.get('books', [])
    if not books:
        return jsonify({'error': '无图书数据'}), 400

    db = get_db()
    inserted = 0
    skipped = 0
    errors = []
    for i, b in enumerate(books):
        try:
            inc, skp = _insert_book(db, b)
            inserted += inc
            skipped += skp
        except Exception as e:
            skipped += 1
            errors.append(f'第{i+1}条：{str(e)}')
    db.commit()
    result = {'success': True, 'inserted': inserted, 'skipped': skipped}
    if errors:
        result['errors'] = errors[:20]
    return jsonify(result)


@app.route('/api/admin/books/import-excel', methods=['POST'])
@admin_required
def import_books_excel():
    """Excel 文件导入"""
    if openpyxl is None:
        return jsonify({'error': '未安装 openpyxl，请运行 pip install openpyxl'}), 500

    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请上传文件'}), 400

    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        return jsonify({'error': '仅支持 .xlsx / .xls 格式'}), 400

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({'error': f'文件无法读取: {str(e)}'}), 400

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # 读取标题行
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return jsonify({'error': '文件为空'}), 400

    headers = [str(h).strip() if h else '' for h in header_row]
    if not any(headers):
        wb.close()
        return jsonify({'error': '未检测到表头'}), 400

    # 列映射
    EXCEL_FIELD_MAP = {
        '完整书名': 'title',
        '正书名': 'title',
        '书名': 'title',
        '副书名': 'subtitle',
        '分卷名': 'volume_name',
        '总分卷名': 'volume_name',
        '分卷号': 'volume_no',
        '总分卷号': 'volume_no',
        '作者': 'author',
        '个人作者': 'author',
        '丛书名': 'series_name',
        '丛书作者': 'series_author',
        '价格': 'price',
        '定价': 'price',
        'ＩＳＢＮ': 'isbn',
        'ISBN': 'isbn',
        '出版社': 'publisher',
        '出版者': 'publisher',
        '出版年': 'pub_year',
        '页数': 'pages',
        '页码': 'pages',
        '主题词': 'keywords',
        '内容简介': 'description',
        '简介': 'description',
        '中图分类': 'clc_number',
        '中图分类号': 'clc_number',
        '分类号': 'clc_number',
        '分类': 'category',
        '类别': 'category',
        '一般附注': 'notes',
        '附注': 'notes',
        '备注': 'notes',
        '封面': 'cover_url',
        '封面地址': 'cover_url',
    }

    col_map = {}  # col_index -> db_field
    for idx, h in enumerate(headers):
        if h in EXCEL_FIELD_MAP:
            col_map[idx] = EXCEL_FIELD_MAP[h]

    if not col_map:
        wb.close()
        return jsonify({'error': f'未能识别任何列。表头: {", ".join(headers[:10])}'}), 400

    db = get_db()
    inserted = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(rows_iter, start=2):
        try:
            if row is None or all(v is None or str(v).strip() == '' for v in row):
                continue
            book = {}
            for col_idx, field in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    if isinstance(val, (int, float)):
                        book[field] = str(val)
                    else:
                        book[field] = str(val).strip()
                else:
                    book[field] = ''

            # 没有书名则跳过
            if not book.get('title', '').strip():
                errors.append(f'第{row_idx}行：无书名，已跳过')
                skipped += 1
                continue

            inc, skp = _insert_book(db, book)
            inserted += inc
            skipped += skp
        except Exception as e:
            skipped += 1
            errors.append(f'第{row_idx}行：{str(e)}')

    db.commit()
    wb.close()

    result = {'success': True, 'inserted': inserted, 'skipped': skipped}
    if errors:
        result['errors'] = errors[:20]
    return jsonify(result)


@app.route('/api/admin/books/template', methods=['GET'])
@admin_required
def download_template():
    """下载 Excel 导入模板"""
    if openpyxl is None:
        return jsonify({'error': '未安装 openpyxl'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '图书导入模板'

    headers = [
        '完整书名', '副书名', '分卷名', '分卷号',
        '作者', '丛书名', '丛书作者',
        '出版社', '出版年', '页数',
        'ＩＳＢＮ', '价格', '分类',
        '主题词', '中图分类',
        '内容简介', '一般附注'
    ]
    ws.append(headers)

    # 示例数据
    ws.append([
        '三体', '', '', '',
        '刘慈欣', '中国科幻基石丛书', '',
        '重庆出版社', '2008', '302',
        '9787536692930', 23.00, '科幻',
        '科幻;外星文明;文革',
        'I247.5',
        '文化大革命如火如荼进行的同时，军方探寻外星文明的绝秘计划红岸工程取得了突破性进展。',
        ''
    ])

    # 样式
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')

    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 列宽
    widths = [20, 15, 12, 10, 15, 20, 15, 18, 10, 8, 18, 10, 10, 25, 12, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='图书导入模板.xlsx'
    )


# ==================== ISBN 查询 ====================
import re
import urllib.request
import urllib.parse
import urllib.error


def _clean_isbn(raw):
    """清洗 ISBN：去除非数字字符，返回纯数字或13位ISBN"""
    s = re.sub(r'[^0-9Xx]', '', raw.strip())
    s = s.upper()
    if len(s) == 10:
        return s
    if len(s) == 13:
        return s
    return s if s else None


def _fetch_alicloud_isbn(isbn_clean):
    """从阿里云 ISBN API 获取图书信息（商业接口，国内可用）"""
    url = f'https://tsisbn.market.alicloudapi.com/isbn/index?isbn={isbn_clean}'
    headers = {
        'Authorization': 'APPCODE ' + ALICLOUD_APPCODE,
        'User-Agent': 'ReaderRecommend/1.0',
    }
    req = urllib.request.Request(url, headers=headers)
    resp = urllib.request.urlopen(req, timeout=10)
    result = json.loads(resp.read().decode('utf-8'))

    # API返回格式：{"code": 1, "msg": "操作成功", "data": {...}}
    if result.get('code') != 1 or not result.get('data'):
        return None

    d = result['data']
    pub_date = (d.get('pubdate') or '').strip()
    pub_year = pub_date[:4] if pub_date else ''

    return {
        '_source': '阿里云ISBN',
        'isbn': isbn_clean,
        'title': (d.get('title') or '').strip(),
        'subtitle': '',
        'author': (d.get('author') or '').strip(),
        'publisher': (d.get('publisher') or '').strip(),
        'pub_year': pub_year,
        'pages': str(d.get('pages', '')).strip(),
        'description': (d.get('summary') or '').strip()[:2000],
        'cover_url': (d.get('img') or '').strip(),
    }


def lookup_isbn(isbn):
    """查询单个 ISBN：先查缓存 → 调阿里云 ISBN API"""
    isbn_clean = _clean_isbn(isbn)
    if not isbn_clean:
        return None, '无效的ISBN号', None

    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row

    # 1) 查缓存
    row = db.execute('SELECT * FROM isbn_cache WHERE isbn = ?', (isbn_clean,)).fetchone()
    if row:
        result = dict(row)
        db.close()
        return result, None, 'cache'

    # 2) 调阿里云 ISBN API
    cache_data = None
    last_error = ''
    try:
        cache_data = _fetch_alicloud_isbn(isbn_clean)
    except urllib.error.URLError as e:
        last_error = f'网络错误: {e.reason}'
    except Exception as e:
        last_error = str(e)

    if not cache_data:
        db.close()
        return None, f'API查询失败：{last_error}', None

    # 3) 存缓存
    source_name = cache_data.get('_source', 'api')
    db.execute('''
        INSERT OR IGNORE INTO isbn_cache (isbn, title, subtitle, author, publisher, pub_year, pages, description, cover_url)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        isbn_clean, cache_data['title'], cache_data.get('subtitle', ''), cache_data.get('author', ''),
        cache_data.get('publisher', ''), cache_data.get('pub_year', ''), cache_data.get('pages', ''),
        cache_data.get('description', ''), cache_data.get('cover_url', '')
    ))
    db.commit()
    db.close()

    return cache_data, None, source_name


@app.route('/api/admin/books/lookup-isbn', methods=['POST'])
@admin_required
def lookup_isbn_api():
    """单个 ISBN 查询"""
    data = request.get_json()
    isbn = (data.get('isbn') or '').strip()
    if not isbn:
        return jsonify({'error': '请提供ISBN号'}), 400

    result, error, source = lookup_isbn(isbn)
    if error:
        return jsonify({'error': error}), 404

    # 清理内部字段
    result.pop('_source', None)
    return jsonify({'success': True, 'book': result, 'source': source})


@app.route('/api/admin/books/lookup-isbn/batch', methods=['POST'])
@admin_required
def lookup_isbn_batch():
    """批量 ISBN 查询并自动导入"""
    data = request.get_json()
    isbns_raw = data.get('isbns', '')
    if isinstance(isbns_raw, str):
        isbns = [i.strip() for i in isbns_raw.replace(',', '\n').split('\n') if i.strip()]
    else:
        isbns = [str(i).strip() for i in isbns_raw if i]

    if not isbns:
        return jsonify({'error': '请提供ISBN号'}), 400

    results = []
    imported = 0
    skipped = 0
    not_found = []

    for isbn in isbns:
        result, error, _source = lookup_isbn(isbn)
        result.pop('_source', None) if result else None
        if error:
            not_found.append({'isbn': isbn, 'error': error})
            continue
        if not result.get('title'):
            not_found.append({'isbn': isbn, 'error': '未找到图书信息'})
            continue

        # 检查是否已在图书库中
        isbn_clean = _clean_isbn(isbn)
        db = get_db()
        exist = db.execute('SELECT id FROM books WHERE isbn = ?', (isbn_clean,)).fetchone()
        if exist:
            skipped += 1
            result['status'] = 'skipped'
        else:
            _insert_book(db, {
                'title': result['title'],
                'subtitle': result.get('subtitle', ''),
                'author': result.get('author', ''),
                'publisher': result.get('publisher', ''),
                'isbn': isbn_clean,
                'pub_year': result.get('pub_year', ''),
                'pages': result.get('pages', ''),
                'description': result.get('description', ''),
                'cover_url': result.get('cover_url', ''),
            })
            db.commit()
            imported += 1
            result['status'] = 'imported'

        results.append(result)

    return jsonify({
        'success': True,
        'total': len(isbns),
        'imported': imported,
        'skipped': skipped,
        'not_found': not_found,
        'results': results,
    })


@app.route('/api/admin/books/<int:book_id>', methods=['DELETE'])
@admin_required
def delete_book(book_id):
    db = get_db()
    db.execute('DELETE FROM click_records WHERE book_id = ?', (book_id,))
    db.execute('DELETE FROM order_items WHERE book_id = ?', (book_id,))
    db.execute('DELETE FROM books WHERE id = ?', (book_id,))
    db.commit()
    return jsonify({'success': True})


@app.route('/api/admin/books/<int:book_id>', methods=['PUT'])
@admin_required
def update_book(book_id):
    db = get_db()
    data = request.get_json()
    sets = []
    vals = []
    for f in BOOK_FIELDS:
        if f in data:
            sets.append(f'{f}=?')
            vals.append(str(data.get(f, '')) if f != 'price' else float(data.get(f, 0)))
    if not sets:
        return jsonify({'error': '无更新数据'}), 400
    sets.append('updated_at=CURRENT_TIMESTAMP')
    vals.append(book_id)
    db.execute(f'UPDATE books SET {",".join(sets)} WHERE id=?', vals)
    db.commit()
    return jsonify({'success': True})


@app.route('/api/admin/recommendations', methods=['GET'])
@admin_required
def admin_recommendations():
    db = get_db()
    threshold = int(get_config('click_threshold') or 10)
    rows = db.execute(
        'SELECT * FROM books WHERE click_count >= ? ORDER BY click_count DESC',
        (threshold,)
    ).fetchall()
    books = []
    for r in rows:
        d = dict(r)
        d['progress'] = min(100, round(d['click_count'] / threshold * 100, 1)) if threshold > 0 else 100
        books.append(d)
    return jsonify({'books': books, 'threshold': threshold, 'total': len(books)})


@app.route('/api/admin/config', methods=['GET'])
@admin_required
def admin_config():
    db = get_db()
    rows = db.execute('SELECT * FROM config').fetchall()
    return jsonify({r['key']: r['value'] for r in rows})


@app.route('/api/admin/config', methods=['PUT'])
@admin_required
def update_config():
    data = request.get_json()
    for key, value in data.items():
        set_config(key, value)
    return jsonify({'success': True})


# ==================== 订单 ====================
@app.route('/api/admin/orders', methods=['GET'])
@admin_required
def admin_orders():
    db = get_db()
    rows = db.execute('SELECT * FROM orders ORDER BY created_at DESC').fetchall()
    orders = []
    for row in rows:
        d = dict(row)
        items = db.execute('''
            SELECT oi.*, b.title, b.author, b.isbn, b.price
            FROM order_items oi JOIN books b ON oi.book_id = b.id
            WHERE oi.order_id = ?
        ''', (row['id'],)).fetchall()
        d['items'] = [dict(i) for i in items]
        orders.append(d)
    return jsonify({'orders': orders})


@app.route('/api/admin/orders', methods=['POST'])
@admin_required
def create_order():
    data = request.get_json()
    book_ids = data.get('book_ids', [])
    vendor = data.get('vendor', '')
    contact = data.get('contact', '')
    note = data.get('note', '')
    if not book_ids:
        return jsonify({'error': '请选择图书'}), 400

    db = get_db()
    total = db.execute(
        'SELECT COALESCE(SUM(price), 0) FROM books WHERE id IN ({})'.format(','.join('?' * len(book_ids))),
        book_ids
    ).fetchone()[0]

    db.execute(
        'INSERT INTO orders (vendor, contact, note, book_ids, total_amount, status) VALUES (?, ?, ?, ?, ?, ?)',
        (vendor, contact, note, json.dumps(book_ids), total, 'submitted')
    )
    order_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

    for bid in book_ids:
        db.execute('INSERT INTO order_items (order_id, book_id, quantity) VALUES (?, ?, 1)', (order_id, bid))
        db.execute("UPDATE books SET status = 'ordered', updated_at = CURRENT_TIMESTAMP WHERE id = ?", (bid,))

    db.commit()
    return jsonify({'success': True, 'order_id': order_id})


@app.route('/api/admin/orders/<int:order_id>/status', methods=['PUT'])
@admin_required
def update_order_status(order_id):
    data = request.get_json()
    new_status = data.get('status', '')
    db = get_db()
    db.execute('UPDATE orders SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (new_status, order_id))
    if new_status == 'received':
        row = db.execute('SELECT book_ids FROM orders WHERE id = ?', (order_id,)).fetchone()
        if row:
            for bid in json.loads(row['book_ids']):
                db.execute("UPDATE books SET status = 'received', updated_at = CURRENT_TIMESTAMP WHERE id = ?", (bid,))
    db.commit()
    return jsonify({'success': True})


# ==================== 统计 ====================
@app.route('/api/admin/stats', methods=['GET'])
@admin_required
def admin_stats():
    db = get_db()
    threshold = int(get_config('click_threshold') or 10)
    return jsonify({
        'total_books': db.execute('SELECT COUNT(*) FROM books').fetchone()[0],
        'recommended': db.execute('SELECT COUNT(*) FROM books WHERE click_count >= ?', (threshold,)).fetchone()[0],
        'ordered': db.execute("SELECT COUNT(*) FROM books WHERE status = 'ordered'").fetchone()[0],
        'received': db.execute("SELECT COUNT(*) FROM books WHERE status = 'received'").fetchone()[0],
        'total_clicks': db.execute('SELECT COALESCE(SUM(click_count), 0) FROM books').fetchone()[0],
        'total_orders': db.execute('SELECT COUNT(*) FROM orders').fetchone()[0],
        'threshold': threshold
    })


# ==================== 智能分类 API ====================
@app.route('/api/admin/classify/<int:book_id>', methods=['POST'])
@admin_required
def classify_single(book_id):
    """对单本图书进行智能分类"""
    db = get_db()
    book = db.execute('SELECT * FROM books WHERE id = ?', (book_id,)).fetchone()
    if not book:
        return jsonify({'error': '图书不存在'}), 404

    use_ai = request.args.get('ai', '0') == '1'
    result = classify_book(
        title=book['title'] or '',
        subtitle=book['subtitle'] or '',
        author=book['author'] or '',
        keywords=book['keywords'] or '',
        series_name=book['series_name'] or '',
        description=book['description'] or '',
        use_ai=use_ai,
    )

    db.execute(
        'UPDATE books SET category=?, clc_number=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
        (result['category'], result['clc_number'], book_id)
    )
    db.commit()

    return jsonify({'success': True, 'book_id': book_id, 'result': result})


@app.route('/api/admin/classify/batch', methods=['POST'])
@admin_required
def classify_batch():
    """批量智能分类（未分类或分类为空的图书）"""
    data = request.get_json() or {}
    book_ids = data.get('book_ids')
    use_ai = request.args.get('ai', '0') == '1'

    db = get_db()
    if book_ids:
        placeholders = ','.join('?' * len(book_ids))
        books = db.execute(
            f'SELECT * FROM books WHERE id IN ({placeholders})', book_ids
        ).fetchall()
    else:
        books = db.execute(
            "SELECT * FROM books WHERE category = '' OR category IS NULL"
        ).fetchall()

    results = []
    for book in books:
        result = classify_book(
            title=book['title'] or '',
            subtitle=book['subtitle'] or '',
            author=book['author'] or '',
            keywords=book['keywords'] or '',
            series_name=book['series_name'] or '',
            description=book['description'] or '',
            use_ai=use_ai,
        )
        db.execute(
            'UPDATE books SET category=?, clc_number=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
            (result['category'], result['clc_number'], book['id'])
        )
        results.append({
            'id': book['id'],
            'title': book['title'],
            'category': result['category'],
            'clc_number': result['clc_number'],
            'confidence': result['confidence'],
            'method': result['method'],
        })
    db.commit()

    return jsonify({'success': True, 'classified': len(results), 'results': results})


@app.route('/api/admin/classify/categories', methods=['GET'])
@admin_required
def list_categories():
    """获取所有支持的学科分类列表（供筛选使用）"""
    from classifier import DISCIPLINE_MAP
    cats = []
    for name, info in DISCIPLINE_MAP.items():
        cat = {
            'name': name,
            'clc': info.get('clc', ''),
            'sub_categories': list(info.get('sub', {}).keys()),
        }
        cats.append(cat)
    return jsonify({'categories': cats})


@app.route('/api/admin/classify/config', methods=['GET', 'POST'])
@admin_required
def classify_config():
    """查看或更新 AI 分类配置"""
    global AI_CLASSIFIER_CONFIG
    if request.method == 'GET':
        return jsonify({
            'enabled': AI_CLASSIFIER_CONFIG.get('enabled', False),
            'api_url': AI_CLASSIFIER_CONFIG.get('api_url', ''),
            'model': AI_CLASSIFIER_CONFIG.get('model', ''),
            'has_api_key': bool(AI_CLASSIFIER_CONFIG.get('api_key', '')),
        })

    data = request.get_json() or {}
    changed = False

    if 'api_url' in data:
        AI_CLASSIFIER_CONFIG['api_url'] = data['api_url']
        changed = True
    if 'api_key' in data:
        AI_CLASSIFIER_CONFIG['api_key'] = data['api_key']
        changed = True
    if 'model' in data:
        AI_CLASSIFIER_CONFIG['model'] = data['model']
        changed = True

    if changed:
        AI_CLASSIFIER_CONFIG['enabled'] = bool(
            AI_CLASSIFIER_CONFIG.get('api_url') and AI_CLASSIFIER_CONFIG.get('api_key')
        )
        try:
            with open(os.path.join(os.path.dirname(__file__), 'ai_config.json'), 'w') as f:
                json.dump({
                    'api_url': AI_CLASSIFIER_CONFIG.get('api_url', ''),
                    'api_key': AI_CLASSIFIER_CONFIG.get('api_key', ''),
                    'model': AI_CLASSIFIER_CONFIG.get('model', ''),
                }, f, indent=2)
        except Exception:
            pass

    return jsonify({'success': True, 'enabled': AI_CLASSIFIER_CONFIG['enabled']})


# ==================== 专业/学科推荐 API ====================
@app.route('/api/admin/recommend-majors/<int:book_id>', methods=['POST'])
@admin_required
def recommend_majors_single(book_id):
    """对单本图书进行专业学科推荐"""
    db = get_db()
    book = db.execute('SELECT * FROM books WHERE id = ?', (book_id,)).fetchone()
    if not book:
        return jsonify({'error': '图书不存在'}), 404

    result = recommend_majors(
        title=book['title'] or '',
        author=book['author'] or '',
        category=book['category'] or '',
        keywords=book['keywords'] or '',
        description=book['description'] or '',
    )

    db.execute(
        'UPDATE books SET recommended_major=?, recommended_subject=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
        (result['recommended_major'], result['recommended_subject'], book_id)
    )
    db.commit()

    return jsonify({'success': True, 'book_id': book_id, 'result': result})


@app.route('/api/admin/recommend-majors/batch', methods=['POST'])
@admin_required
def recommend_majors_batch():
    """批量为未设置专业推荐的图书进行推荐"""
    db = get_db()
    books = db.execute(
        "SELECT * FROM books WHERE recommended_major = '' OR recommended_major IS NULL"
    ).fetchall()

    results = []
    for book in books:
        result = recommend_majors(
            title=book['title'] or '',
            author=book['author'] or '',
            category=book['category'] or '',
            keywords=book['keywords'] or '',
            description=book['description'] or '',
        )
        db.execute(
            'UPDATE books SET recommended_major=?, recommended_subject=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
            (result['recommended_major'], result['recommended_subject'], book['id'])
        )
        results.append({
            'id': book['id'],
            'title': book['title'],
            'recommended_major': result['recommended_major'],
            'recommended_subject': result['recommended_subject'],
            'method': result['method'],
        })
    db.commit()

    ai_count = sum(1 for r in results if r['method'] == 'ai')
    rule_count = sum(1 for r in results if r['method'] == 'rule')
    return jsonify({
        'success': True,
        'total': len(results),
        'ai_count': ai_count,
        'rule_count': rule_count,
        'results': results,
    })


# ==================== 静态文件 ====================
def _no_cache_response(filepath, mimetype='text/html'):
    """发送带禁用缓存的静态文件"""
    from flask import make_response
    response = make_response(send_from_directory(app.static_folder, filepath, mimetype=mimetype))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/')
def index():
    return _no_cache_response('index.html')


@app.route('/admin')
def admin():
    return _no_cache_response('admin.html')


@app.route('/admin/<path:filename>')
def admin_static(filename):
    return send_from_directory(app.static_folder, filename)


# ==================== 启动 ====================
# 自动初始化数据库（gunicorn 导入 app 时也会执行）
init_db()


if __name__ == '__main__':
    import sys as _sys
    _sys.stdout = io.TextIOWrapper(_sys.stdout.buffer, encoding='utf-8')
    print("=" * 50)
    print("[Reader Recommend] 读者荐读系统 已启动")
    print("   读者端: http://localhost:5000")
    print("   管理端: http://localhost:5000/admin")
    print(f"   管理员密码: {ADMIN_PASSWORD}")
    print("   Excel模板: 管理后台 -> 图书管理 -> 下载模板")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5000, debug=False)
